import csv
import json
import random
import traceback
from datetime import timedelta
from email.mime.image import MIMEImage
from io import BytesIO
from pathlib import Path
from functools import wraps

from django.contrib.auth.hashers import check_password, make_password
from django.core.cache import cache
from django.core.mail import EmailMultiAlternatives, send_mail
from django.core import signing
from django.db import transaction
from django.db.models import Count
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Administrador, Aprendiz, AprendizGaes, Entregable, Evaluacion, Ficha, Gaes, Instructor, Proyecto, Rol, Trimestre, Usuario
from .services.gemini_sdk import GeminiSdkError, generar_respuesta_gemini

# ============================================================
# CONFIGURACION Y UTILIDADES BASE
# ============================================================

TOKEN_MAX_AGE = 60 * 60 * 24
PASSWORD_RESET_CODE_TTL = 60 * 3


def ensure_default_roles():
    """Crea los roles base si todavia no existen."""
    defaults = {1: 'administrador', 2: 'instructor', 3: 'aprendiz'}
    for role_id, nombre in defaults.items():
        Rol.objects.update_or_create(id=role_id, defaults={'nombre_rol': nombre})


def normalize_text(value):
    """Normaliza texto recibido desde formularios o JSON."""
    return (value or '').strip()


def is_strong_password(value):
    """Valida una contraseña con reglas basicas de seguridad."""
    value = value or ''
    return (
        len(value) >= 8
        and any(ch.isupper() for ch in value)
        and any(ch.islower() for ch in value)
        and any(ch.isdigit() for ch in value)
        and any(not ch.isalnum() for ch in value)
    )


def choice_to_db(value):
    """Normaliza enums del frontend al formato almacenado en MySQL."""
    return normalize_text(value).lower()


def choice_to_api(value):
    """Normaliza enums de MySQL al formato esperado por el frontend."""
    return normalize_text(value).upper()


def parse_json(request):
    """Lee el body JSON de la request de forma segura."""
    try:
        return json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return {}


def get_payload_id(data, key):
    """Obtiene ids desde payload plano o desde objetos anidados."""
    value = data.get(key)
    if value not in (None, ''):
        return value
    if key.endswith('Id'):
        nested = data.get(key[:-2])
        if isinstance(nested, dict):
            nested_value = nested.get('id')
            if nested_value not in (None, ''):
                return nested_value
    return None


def validate_unique_usuario_fields(correo, numero_documento, exclude_user_id=None):
    """Valida que correo y documento no esten duplicados en otro usuario."""
    correo = normalize_text(correo).lower()
    numero_documento = normalize_text(numero_documento)

    correo_qs = Usuario.objects.filter(correo=correo)
    documento_qs = Usuario.objects.filter(numero_documento=numero_documento)

    if exclude_user_id is not None:
        correo_qs = correo_qs.exclude(id=exclude_user_id)
        documento_qs = documento_qs.exclude(id=exclude_user_id)

    if correo and correo_qs.exists():
        return 'Este correo ya esta en uso'
    if numero_documento and documento_qs.exists():
        return 'Este numero de documento ya esta en uso'
    return None


def validate_aprendiz_en_ficha(usuario, ficha, exclude_aprendiz_id=None):
    """Evita duplicar el mismo aprendiz dentro de una misma ficha."""
    if not usuario or not ficha:
        return None

    queryset = Aprendiz.objects.select_related('usuario', 'ficha').filter(ficha=ficha)
    if exclude_aprendiz_id is not None:
        queryset = queryset.exclude(id=exclude_aprendiz_id)

    if queryset.filter(usuario=usuario).exists():
        return 'Este aprendiz ya esta registrado en esa ficha'

    if queryset.filter(usuario__numero_documento=usuario.numero_documento).exists():
        return 'Ya existe un aprendiz con ese documento en esa ficha'

    if queryset.filter(usuario__nombre__iexact=usuario.nombre, usuario__apellido__iexact=usuario.apellido).exists():
        return 'Ya existe un aprendiz con ese mismo nombre en esa ficha'

    return None


def issue_token(usuario):
    """Genera un token firmado simple para autenticacion del frontend."""
    return signing.dumps(
        {'user_id': usuario.id, 'rol': usuario.rol.nombre_rol if usuario.rol else '', 'correo': usuario.correo},
        salt='mapeo-auth',
    )


def send_password_change_email(usuario):
    """Intenta enviar un correo de cambio obligatorio de contraseña para instructores."""
    subject = 'Cambio obligatorio de contraseña - GesPro'
    text_message = (
        f'Hola {usuario.nombre},\n\n'
        'Tu cuenta de instructor fue creada por un administrador y debes cambiar tu contraseña temporal.\n'
        'Ingresa al sistema, usa la opción "¿Olvidaste tu contraseña?" y registra tu nueva clave.\n\n'
        f'Correo: {usuario.correo}\n'
        f'Documento: {usuario.numero_documento}\n\n'
        'Cuando la actualices, esa será la contraseña que usarás en adelante.\n\n'
        'Equipo GesPro.'
    )
    html_message = render_to_string(
        'emails/instructor_password_change.html',
        {
            'usuario': usuario,
        },
    )

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[usuario.correo],
    )
    email.attach_alternative(html_message, 'text/html')

    logo_path = Path(settings.BASE_DIR) / 'frontend_assets' / 'img' / 'logo3.png'
    if logo_path.exists():
        with logo_path.open('rb') as image_file:
            logo = MIMEImage(image_file.read())
            logo.add_header('Content-ID', '<logo3>')
            logo.add_header('Content-Disposition', 'inline', filename=logo_path.name)
            email.attach(logo)

    try:
        email.send(fail_silently=False)
        return True
    except Exception:
        return False


def get_password_reset_cache_key(correo):
    """Construye la llave de cache para codigos de recuperacion."""
    return f'password-reset:{correo}'


def send_password_reset_code_email(usuario, codigo):
    """Envia al usuario un codigo temporal para recuperar su contraseña."""
    minutos = PASSWORD_RESET_CODE_TTL // 60
    expires_at = timezone.localtime(timezone.now() + timedelta(minutes=minutos)).strftime('%d/%m/%Y %I:%M %p')
    subject = 'Código de recuperación - GesPro'
    text_message = (
        f'Hola {usuario.nombre},\n\n'
        'Recibimos una solicitud para restablecer tu contraseña en GesPro.\n'
        f'Tu código de recuperación es: {codigo}\n'
        f'Este código tiene una validez de {minutos} minutos.\n'
        f'Fecha límite aproximada: {expires_at}\n\n'
        'Si no solicitaste este cambio, puedes ignorar este correo.\n'
        'Equipo GesPro.'
    )
    html_message = render_to_string(
        'emails/password_reset_code.html',
        {
            'usuario': usuario,
            'codigo': codigo,
            'minutes': minutos,
            'expires_at': expires_at,
        },
    )

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[usuario.correo],
    )
    email.attach_alternative(html_message, 'text/html')

    logo_path = Path(settings.BASE_DIR) / 'frontend_assets' / 'img' / 'logo3.png'
    if logo_path.exists():
        with logo_path.open('rb') as image_file:
            logo = MIMEImage(image_file.read())
            logo.add_header('Content-ID', '<logo3>')
            logo.add_header('Content-Disposition', 'inline', filename=logo_path.name)
            email.attach(logo)

    email.send(fail_silently=False)


def get_request_user(request):
    """Obtiene el usuario autenticado desde el header Authorization."""
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return None
    token = auth.split(' ', 1)[1]
    try:
        payload = signing.loads(token, salt='mapeo-auth', max_age=TOKEN_MAX_AGE)
    except signing.BadSignature:
        return None
    return Usuario.objects.filter(id=payload.get('user_id')).select_related('rol').first()


def api_login_required(view):
    """Protege endpoints API que requieren autenticacion."""
    @wraps(view)
    def wrapper(request, *args, **kwargs):
        if get_request_user(request) is None:
            return JsonResponse({'detail': 'Token invalido o ausente'}, status=401)
        return view(request, *args, **kwargs)

    return wrapper


# ============================================================
# PAGINAS PRINCIPALES Y DASHBOARDS
# ============================================================

def page_index(request):
    """Renderiza la pagina principal."""
    return render(request, 'index.html')


def page_login(request):
    """Renderiza la pantalla de login."""
    ensure_default_roles()
    return render(request, 'login.html')


def page_forgot_password(request):
    """Renderiza la pantalla de recuperacion de contrasena."""
    return render(request, 'forgot-password.html', {'prefill_correo': request.GET.get('correo', '')})


def page_register(request):
    """Renderiza la pantalla de registro."""
    ensure_default_roles()
    return render(request, 'register.html')


def build_auth_success_response(usuario):
    """Construye la respuesta estandar de autenticacion."""
    rol_nombre = usuario.rol.nombre_rol.lower() if usuario.rol else ''

    if rol_nombre == 'instructor' and (usuario.debe_cambiar_password or usuario.password_temporal):
        correo_enviado = send_password_change_email(usuario)
        return JsonResponse(
            {
                'requiresPasswordChange': True,
                'message': 'Tu cuenta de instructor debe cambiar la contrasena antes de ingresar.',
                'emailSent': correo_enviado,
                'correo': usuario.correo,
            },
            status=403,
        )

    dashboard = {
        'aprendiz': '/dashboard/aprendiz',
        'instructor': '/dashboard/instructor',
        'administrador': '/dashboard/administrador',
    }.get(rol_nombre, '/')

    return JsonResponse(
        {
            'token': issue_token(usuario),
            'usuarioId': usuario.id,
            'nombre': usuario.nombre,
            'apellido': usuario.apellido,
            'correo': usuario.correo,
            'rol': rol_nombre,
            'dashboard': dashboard,
        }
    )


def page_dashboard_administrador(request):
    """Renderiza el dashboard del administrador."""
    return render(request, 'administrador-dashboard.html')


@csrf_exempt
@require_http_methods(['GET', 'POST'])
def asistente(request):
    """Asistente IA basado en Gemini para preguntas en texto."""
    respuesta = ''
    error = ''

    if request.method == 'POST':
        if get_request_user(request) is None:
            return JsonResponse({'detail': 'Token invalido o ausente'}, status=401)
        data = parse_json(request) if request.content_type == 'application/json' else request.POST
        texto = normalize_text(data.get('texto'))
        try:
            respuesta = generar_respuesta_gemini(texto)
        except GeminiSdkError as exc:
            error = str(exc)

        if request.content_type == 'application/json':
            if error:
                return JsonResponse({'detail': error}, status=400)
            return JsonResponse({'respuesta': respuesta})

    return render(request, 'asistente.html', {'respuesta': respuesta, 'error': error})


def page_dashboard_aprendiz(request):
    """Renderiza el dashboard del aprendiz."""
    return render(request, 'aprendiz-dashboard.html')


def page_dashboard_instructor(request):
    """Renderiza el dashboard del instructor."""
    return render(request, 'instructor-dashboard.html')




# ============================================================
# SERIALIZADORES DE MODELOS A JSON
# ============================================================

def serialize_rol(rol):
    """Convierte un rol a JSON compatible con el frontend."""
    return {'id': rol.id, 'nombreRol': rol.nombre_rol}


def serialize_usuario(usuario):
    """Convierte un usuario a JSON compatible con el frontend."""
    return {
        'id': usuario.id,
        'nombre': usuario.nombre,
        'apellido': usuario.apellido,
        'correo': usuario.correo,
        'tipoDocumento': usuario.tipo_documento,
        'numeroDocumento': usuario.numero_documento,
        'estado': choice_to_api(usuario.estado),
        'rol': serialize_rol(usuario.rol) if usuario.rol else None,
        'fotoPerfil': '',
        'fotoUrl': '',
        'debeCambiarPassword': bool(usuario.debe_cambiar_password),
        'passwordTemporal': bool(usuario.password_temporal),
    }


def serialize_ficha(ficha):
    """Convierte una ficha a JSON compatible con el frontend."""
    return {
        'id': ficha.id,
        'codigoFicha': ficha.codigo_ficha,
        'codigo': ficha.codigo_ficha,
        'programaFormacion': ficha.programa_formacion,
        'nivel': choice_to_api(ficha.nivel),
        'jornada': choice_to_api(ficha.jornada),
        'modalidad': choice_to_api(ficha.modalidad),
        'fechaInicio': ficha.fecha_inicio.isoformat() if ficha.fecha_inicio else None,
        'fechaFin': ficha.fecha_fin.isoformat() if ficha.fecha_fin else None,
        'estado': choice_to_api(ficha.estado),
    }


def serialize_gaes(gaes, include_integrantes=False):
    """Convierte un GAES a JSON y opcionalmente incluye sus integrantes."""
    data = {'id': gaes.id, 'nombre': gaes.nombre, 'fichaId': gaes.ficha_id, 'estado': 'ACTIVO'}
    if include_integrantes:
        aprendiz_ids = AprendizGaes.objects.filter(gaes=gaes).values_list('aprendiz_id', flat=True)
        aprendices = Aprendiz.objects.filter(id__in=aprendiz_ids).select_related('usuario', 'ficha')
        data['integrantes'] = [serialize_aprendiz(item, nested=True) for item in aprendices]
    return data


def serialize_trimestre(trimestre):
    """Convierte un trimestre a JSON compatible con el frontend."""
    return {
        'id': trimestre.id,
        'numero': trimestre.numero,
        'fichaId': trimestre.ficha_id,
        'ficha': serialize_ficha(trimestre.ficha) if trimestre.ficha_id else None,
        'fechaInicio': trimestre.fecha_inicio.isoformat() if trimestre.fecha_inicio else None,
        'fechaFin': trimestre.fecha_fin.isoformat() if trimestre.fecha_fin else None,
        'estado': choice_to_api(trimestre.estado),
    }


def serialize_aprendiz(aprendiz, nested=False):
    """Convierte un aprendiz a JSON; nested evita objetos anidados grandes."""
    gaes_link = AprendizGaes.objects.filter(aprendiz=aprendiz).select_related('gaes').first()
    gaes = gaes_link.gaes if gaes_link else None
    data = {
        'id': aprendiz.id,
        'usuarioId': aprendiz.usuario_id,
        'fichaId': aprendiz.ficha_id,
        'gaesId': gaes.id if gaes else None,
        'gaesNombre': gaes.nombre if gaes else None,
        'usuarioNombre': aprendiz.usuario.nombre,
        'usuarioApellido': aprendiz.usuario.apellido,
        'usuarioCorreo': aprendiz.usuario.correo,
        'numeroDocumento': aprendiz.usuario.numero_documento,
        'fichaCodigoFicha': aprendiz.ficha.codigo_ficha if aprendiz.ficha_id else None,
        'fichaProgramaFormacion': aprendiz.ficha.programa_formacion if aprendiz.ficha_id else None,
        'nivel': choice_to_api(aprendiz.ficha.nivel) if aprendiz.ficha_id else None,
        'esLider': False,
        'estado': choice_to_api(aprendiz.usuario.estado),
    }
    if not nested:
        data['usuario'] = serialize_usuario(aprendiz.usuario)
        data['ficha'] = serialize_ficha(aprendiz.ficha) if aprendiz.ficha_id else None
        data['gaes'] = serialize_gaes(gaes) if gaes else None
    return data


def serialize_instructor(instructor):
    """Convierte un instructor a JSON compatible con el frontend."""
    return {
        'id': instructor.id,
        'especialidad': instructor.especialidad,
        'estado': choice_to_api(instructor.usuario.estado),
        'usuarioId': instructor.usuario_id,
        'fichaId': instructor.ficha_id,
        'usuario': serialize_usuario(instructor.usuario),
        'ficha': serialize_ficha(instructor.ficha) if instructor.ficha_id else None,
    }


def get_preferred_instructor_by_user_id(usuario_id):
    """Obtiene el perfil instructor mas util cuando existen duplicados para un usuario."""
    queryset = Instructor.objects.select_related('usuario', 'ficha').filter(usuario_id=usuario_id).order_by('-id')
    return queryset.filter(ficha__isnull=False).first() or queryset.first()


def serialize_administrador(item):
    """Convierte un administrador a JSON."""
    return {'id': item.id, 'usuario': serialize_usuario(item.usuario)}


def cantidad_trimestres_por_nivel(nivel):
    nivel_normalizado = (nivel or '').strip().lower()
    return 7 if nivel_normalizado == Ficha.Nivel.TECNOLOGO else 4


def sincronizar_trimestres_ficha(ficha):
    if not ficha or not ficha.fecha_inicio or not ficha.fecha_fin:
        return

    cantidad = cantidad_trimestres_por_nivel(ficha.nivel)
    existentes = list(Trimestre.objects.filter(ficha=ficha).order_by('numero'))
    existentes_por_numero = {item.numero: item for item in existentes}

    if len(existentes) >= cantidad:
        return

    dias_totales = max((ficha.fecha_fin - ficha.fecha_inicio).days + 1, cantidad)
    dias_por_trimestre = max(dias_totales // cantidad, 1)

    for numero in range(1, cantidad + 1):
        if numero in existentes_por_numero:
            continue

        fecha_inicio = ficha.fecha_inicio + timedelta(days=(numero - 1) * dias_por_trimestre)
        fecha_fin = fecha_inicio + timedelta(days=dias_por_trimestre - 1)
        if numero == cantidad or fecha_fin > ficha.fecha_fin:
            fecha_fin = ficha.fecha_fin

        Trimestre.objects.create(
            numero=numero,
            ficha=ficha,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estado=Trimestre.Estado.ACTIVO,
        )


def serialize_proyecto(proyecto):
    """Convierte un proyecto a JSON compatible con el frontend."""
    return {
        'id': proyecto.id,
        'nombre': proyecto.nombre,
        'descripcion': proyecto.descripcion,
        'gaesId': proyecto.gaes_id,
        'aprendizLiderId': None,
        'trimestre': None,
        'trimestreId': None,
        'documentoInicial': '',
        'estado': choice_to_api(proyecto.estado),
        'fechaCreacion': None,
        'fechaActualizacion': None,
        'fechaInicio': proyecto.fecha_inicio.isoformat() if proyecto.fecha_inicio else None,
        'fechaFin': proyecto.fecha_fin.isoformat() if proyecto.fecha_fin else None,
    }


def proyecto_estado_to_db(value):
    estado = normalize_text(value).upper()
    mapping = {
        'EN_PROCESO': Proyecto.Estado.EN_PROCESO,
        'EN_DESARROLLO': Proyecto.Estado.EN_PROCESO,
        'EN_REVISION': Proyecto.Estado.EN_PROCESO,
        'APROBADO': Proyecto.Estado.FINALIZADO,
        'FINALIZADO': Proyecto.Estado.FINALIZADO,
        'CANCELADO': Proyecto.Estado.CANCELADO,
        'RECHAZADO': Proyecto.Estado.CANCELADO,
    }
    return mapping.get(estado, Proyecto.Estado.EN_PROCESO)


def serialize_evaluacion(item):
    """Convierte una evaluacion a JSON compatible con el frontend."""
    return {
        'id': item.id,
        'aprendizId': item.aprendiz_id,
        'gaesId': item.gaes_id,
        'evaluadorId': item.evaluador_id,
        'entregableId': item.entregable_id,
        'aprendiz': serialize_aprendiz(item.aprendiz) if item.aprendiz_id else None,
        'gaes': serialize_gaes(item.gaes) if item.gaes_id else None,
        'evaluador': serialize_instructor(item.evaluador),
        'calificacion': float(item.calificacion) if item.calificacion is not None else None,
        'observaciones': item.observaciones,
        'fecha': item.fecha.isoformat() if item.fecha else None,
        'tipo': 'GAES' if item.gaes_id else 'INDIVIDUAL',
        'entregableNombre': item.entregable.nombre if item.entregable_id else '',
    }


def serialize_entregable(item):
    """Convierte un entregable a JSON compatible con el frontend."""
    return {
        'id': item.id,
        'nombre': item.nombre,
        'descripcion': item.descripcion,
        'proyectoId': item.proyecto_id,
        'trimestreId': item.trimestre_id,
        'aprendizId': item.aprendiz_id,
        'url': item.url or '',
        'archivo': bool(item.archivo),
        'nombreArchivo': item.nombre_archivo or '',
    }


def get_fk(model, value):
    """Busca una llave foranea opcional por id."""
    return model.objects.filter(id=value).first() if value else None


def normalize_date(value):
    """Convierte fechas tipo string YYYY-MM-DD a objeto date."""
    if not value:
        return None
    if hasattr(value, 'isoformat'):
        return value
    return parse_date(str(value))




def csv_response(filename, headers, rows):
    """Genera una descarga CSV simple para reportes."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def excel_response(filename, headers, rows, sheet_name='Reporte'):
    """Genera un archivo Excel real en formato XLSX."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]

    header_fill = PatternFill(fill_type='solid', fgColor='355070')
    header_font = Font(color='FFFFFF', bold=True)
    centered = Alignment(horizontal='center', vertical='center')

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered

    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical='center')

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 35)

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def pdf_response(filename, headers, rows, title='Reporte'):
    """Genera un PDF tabular con logos usando ReportLab."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.textColor = colors.HexColor('#2f4867')
    title_style.fontSize = 22
    title_style.leading = 26

    subtitle_style = styles['Normal']
    subtitle_style.textColor = colors.HexColor('#41556d')
    subtitle_style.fontSize = 10

    story = []
    img_dir = Path(__file__).resolve().parent.parent / 'frontend_assets' / 'img'
    logo_sena = img_dir / 'logo_sena.png'
    logo_gestion = img_dir / 'logo3.png'

    logo_cells = []
    if logo_sena.exists():
        logo_cells.append(Image(str(logo_sena), width=22 * mm, height=22 * mm))
    else:
        logo_cells.append('')

    logo_cells.append(
        Paragraph(
            f'<b>{title}</b><br/><font size="10">SENA | Sistema de Gestion de Proyectos Formativos</font>',
            title_style,
        )
    )

    if logo_gestion.exists():
        logo_cells.append(Image(str(logo_gestion), width=30 * mm, height=18 * mm))
    else:
        logo_cells.append('')

    header_table = Table([logo_cells], colWidths=[28 * mm, 190 * mm, 38 * mm])
    header_table.setStyle(
        TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ])
    )
    story.append(header_table)
    story.append(Spacer(1, 4 * mm))

    table_data = [list(headers)] + [list(row) for row in rows]
    available_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    if len(headers) == 8:
        col_widths = [12 * mm, 24 * mm, 26 * mm, 58 * mm, 26 * mm, 34 * mm, 32 * mm, 24 * mm]
    else:
        col_widths = [available_width / max(len(headers), 1)] * len(headers)

    report_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    report_table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#355070')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7f9fc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#eef3f9'), colors.white]),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#233142')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c6d1df')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ])
    )
    story.append(report_table)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph('Generado desde Django', subtitle_style))

    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# AUTENTICACION Y PERFIL ACTUAL
# ============================================================

@csrf_exempt
@require_http_methods(['POST'])
def auth_register(request):
    """Registra un usuario y crea su perfil base segun el rol."""
    ensure_default_roles()
    data = parse_json(request)
    validation_error = validate_unique_usuario_fields(data.get('correo'), data.get('numeroDocumento'))
    if validation_error:
        return JsonResponse(validation_error, safe=False, status=400)

    rol = get_object_or_404(Rol, id=data.get('rolId'))
    if rol.nombre_rol.lower() != 'aprendiz':
        return JsonResponse('El registro publico solo esta disponible para aprendices', safe=False, status=403)
    if not is_strong_password(data.get('password', '')):
        return JsonResponse('La contrasena no cumple con los requisitos de seguridad', safe=False, status=400)

    usuario = Usuario.objects.create(
        nombre=normalize_text(data.get('nombre')),
        apellido=normalize_text(data.get('apellido')),
        correo=normalize_text(data.get('correo')).lower(),
        tipo_documento=normalize_text(data.get('tipoDocumento')),
        numero_documento=normalize_text(data.get('numeroDocumento')),
        estado=Usuario.Estado.ACTIVO,
        rol=rol,
        password=make_password(data.get('password', '')),
        debe_cambiar_password=False,
        password_temporal=False,
    )
    ficha = Ficha.objects.order_by('id').first()
    if ficha is None:
        return JsonResponse('No hay fichas registradas para crear un aprendiz', safe=False, status=400)
    Aprendiz.objects.get_or_create(usuario=usuario, defaults={'ficha': ficha})
    return JsonResponse('Usuario registrado exitosamente', safe=False, status=201)


@csrf_exempt
@require_http_methods(['POST'])
def auth_login(request):
    """Inicia sesion y devuelve token + dashboard segun rol."""
    ensure_default_roles()
    data = parse_json(request)
    usuario = Usuario.objects.filter(correo=normalize_text(data.get('correo')).lower()).select_related('rol').first()
    if not usuario:
        return JsonResponse('Usuario no encontrado', safe=False, status=401)

    if not check_password(data.get('password', ''), usuario.password):
        return JsonResponse('Contrasena incorrecta', safe=False, status=401)
    return build_auth_success_response(usuario)


@csrf_exempt
@require_http_methods(['POST'])
def auth_request_password_code(request):
    """Genera y envia un codigo de recuperacion al correo del usuario."""
    data = parse_json(request)
    correo = normalize_text(data.get('correo')).lower()
    numero_documento = normalize_text(data.get('numeroDocumento'))

    usuario = Usuario.objects.filter(correo=correo).first()
    if not usuario:
        return JsonResponse('No existe un usuario con ese correo', safe=False, status=404)
    if usuario.numero_documento != numero_documento:
        return JsonResponse('El numero de documento no coincide con el usuario', safe=False, status=400)

    codigo = f'{random.randint(0, 999999):06d}'
    cache.set(
        get_password_reset_cache_key(correo),
        {'usuario_id': usuario.id, 'numero_documento': numero_documento, 'codigo': codigo},
        PASSWORD_RESET_CODE_TTL,
    )

    try:
        send_password_reset_code_email(usuario, codigo)
    except Exception as exc:
        traceback.print_exc()
        return JsonResponse(
            {
                'message': 'No se pudo enviar el codigo de recuperacion al correo',
                'detail': str(exc),
            },
            status=500,
        )

    return JsonResponse({'message': 'Codigo de recuperacion enviado al correo'})


@csrf_exempt
@require_http_methods(['POST'])
def auth_forgot_password(request):
    """Permite restablecer la contraseña validando correo y documento."""
    data = parse_json(request)
    correo = normalize_text(data.get('correo')).lower()
    numero_documento = normalize_text(data.get('numeroDocumento'))
    codigo = normalize_text(data.get('codigo'))
    nueva_password = data.get('nuevaPassword') or ''

    usuario = Usuario.objects.filter(correo=correo).first()
    if not usuario:
        return JsonResponse('No existe un usuario con ese correo', safe=False, status=404)
    if usuario.numero_documento != numero_documento:
        return JsonResponse('El numero de documento no coincide con el usuario', safe=False, status=400)
    cache_data = cache.get(get_password_reset_cache_key(correo))
    if not cache_data:
        return JsonResponse('Primero debes solicitar el codigo de recuperacion', safe=False, status=400)
    if cache_data.get('usuario_id') != usuario.id or cache_data.get('numero_documento') != numero_documento:
        return JsonResponse('El codigo de recuperacion ya no es valido para este usuario', safe=False, status=400)
    if cache_data.get('codigo') != codigo:
        return JsonResponse('El codigo de recuperacion es incorrecto o ya vencio', safe=False, status=400)
    if not is_strong_password(nueva_password):
        return JsonResponse('La nueva contrasena no cumple con los requisitos de seguridad', safe=False, status=400)

    usuario.password = make_password(nueva_password)
    usuario.debe_cambiar_password = False
    usuario.password_temporal = False
    usuario.save(update_fields=['password', 'debe_cambiar_password', 'password_temporal'])
    cache.delete(get_password_reset_cache_key(correo))
    return JsonResponse({'message': 'Contrasena actualizada exitosamente'})


@csrf_exempt
@require_http_methods(['POST'])
def auth_validate_token(request):
    """Valida el token firmado enviado por el frontend."""
    usuario = get_request_user(request)
    if usuario is None:
        return JsonResponse('Token expirado o invalido', safe=False, status=401)
    return JsonResponse({'usuarioId': usuario.id, 'rol': usuario.rol.nombre_rol.lower() if usuario.rol else ''})


@csrf_exempt
@require_http_methods(['POST'])
def auth_logout(request):
    """Confirma logout; la limpieza real la hace el frontend."""
    return JsonResponse({'message': 'Logout exitoso', 'status': 'success'})


@api_login_required
def usuario_actual(request):
    """Devuelve el usuario autenticado actual."""
    return JsonResponse(serialize_usuario(get_request_user(request)))


# ============================================================
# API ROLES
# ============================================================

@csrf_exempt
@require_http_methods(['GET', 'POST'])
def roles_collection(request):
    """GET lista roles, POST crea un rol nuevo."""
    ensure_default_roles()
    if request.method == 'GET':
        return JsonResponse([serialize_rol(item) for item in Rol.objects.all()], safe=False)
    data = parse_json(request)
    rol = Rol.objects.create(nombre_rol=normalize_text(data.get('nombreRol')))
    return JsonResponse(serialize_rol(rol), status=201)


@csrf_exempt
@require_http_methods(['PUT', 'DELETE'])
def roles_detail(request, item_id):
    """PUT actualiza un rol, DELETE elimina un rol."""
    item = get_object_or_404(Rol, id=item_id)
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    item.nombre_rol = normalize_text(data.get('nombreRol'))
    item.save()
    return JsonResponse(serialize_rol(item))


# ============================================================
# API USUARIOS
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def usuarios_collection(request):
    """GET lista usuarios, POST crea un usuario."""
    if request.method == 'GET':
        queryset = Usuario.objects.select_related('rol').all()
        rol_query = normalize_text(request.GET.get('rol'))
        if rol_query:
            nombres = [part.strip().lower() for part in rol_query.split(',') if part.strip()]
            queryset = queryset.filter(rol__nombre_rol__in=nombres)
        return JsonResponse([serialize_usuario(item) for item in queryset], safe=False)

    data = parse_json(request)
    validation_error = validate_unique_usuario_fields(data.get('correo'), data.get('numeroDocumento'))
    if validation_error:
        return JsonResponse(validation_error, safe=False, status=400)
    rol = get_fk(Rol, (data.get('rol') or {}).get('id') or data.get('rolId'))
    rol_nombre = rol.nombre_rol.lower() if rol else ''
    password_raw = data.get('password') or ('Temporal123*' if rol_nombre == 'instructor' else 'Temporal123*')
    usuario = Usuario.objects.create(
        nombre=normalize_text(data.get('nombre')),
        apellido=normalize_text(data.get('apellido')),
        correo=normalize_text(data.get('correo')).lower(),
        tipo_documento=normalize_text(data.get('tipoDocumento')),
        numero_documento=normalize_text(data.get('numeroDocumento')),
        estado=choice_to_db(data.get('estado')) or Usuario.Estado.ACTIVO,
        rol=rol,
        password=make_password(password_raw),
        debe_cambiar_password=rol_nombre == 'instructor',
        password_temporal=rol_nombre == 'instructor',
    )
    if rol_nombre == 'instructor':
        Instructor.objects.get_or_create(usuario=usuario)
    elif rol_nombre == 'administrador':
        Administrador.objects.get_or_create(usuario=usuario)
    elif rol_nombre == 'aprendiz':
        ficha = Ficha.objects.order_by('id').first()
        if ficha:
            Aprendiz.objects.get_or_create(usuario=usuario, defaults={'ficha': ficha})
    return JsonResponse(serialize_usuario(usuario), status=201)


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def usuarios_detail(request, item_id):
    """GET, PUT o DELETE sobre un usuario especifico."""
    item = get_object_or_404(Usuario.objects.select_related('rol'), id=item_id)
    if request.method == 'GET':
        return JsonResponse(serialize_usuario(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)

    data = parse_json(request)
    validation_error = validate_unique_usuario_fields(
        data.get('correo') or item.correo,
        data.get('numeroDocumento') or item.numero_documento,
        exclude_user_id=item.id,
    )
    if validation_error:
        return JsonResponse(validation_error, safe=False, status=400)
    item.nombre = normalize_text(data.get('nombre')) or item.nombre
    item.apellido = normalize_text(data.get('apellido')) or item.apellido
    item.correo = normalize_text(data.get('correo')).lower() or item.correo
    item.tipo_documento = normalize_text(data.get('tipoDocumento'))
    item.numero_documento = normalize_text(data.get('numeroDocumento')) or item.numero_documento
    item.estado = choice_to_db(data.get('estado')) or item.estado
    rol = get_fk(Rol, (data.get('rol') or {}).get('id') or data.get('rolId'))
    if rol:
        item.rol = rol
    if data.get('password'):
        item.password = make_password(data.get('password'))
        item.debe_cambiar_password = False
        item.password_temporal = False
    rol_nombre = item.rol.nombre_rol.lower() if item.rol else ''
    if rol_nombre == 'instructor' and not data.get('password'):
        item.debe_cambiar_password = True
        item.password_temporal = True
    elif rol_nombre != 'instructor' and not data.get('password'):
        item.debe_cambiar_password = False
        item.password_temporal = False
    item.save()
    if rol_nombre == 'instructor':
        Instructor.objects.get_or_create(usuario=item)
    elif rol_nombre == 'administrador':
        Administrador.objects.get_or_create(usuario=item)
    return JsonResponse(serialize_usuario(item))


@csrf_exempt
@api_login_required
@require_http_methods(['PATCH'])
def usuario_activar(request, item_id):
    """Activa un usuario sin eliminarlo."""
    item = get_object_or_404(Usuario, id=item_id)
    item.estado = Usuario.Estado.ACTIVO
    item.save(update_fields=['estado'])
    return JsonResponse(serialize_usuario(item))


@csrf_exempt
@api_login_required
@require_http_methods(['PATCH'])
def usuario_desactivar(request, item_id):
    """Desactiva un usuario sin eliminarlo."""
    item = get_object_or_404(Usuario, id=item_id)
    item.estado = Usuario.Estado.INACTIVO
    item.save(update_fields=['estado'])
    return JsonResponse(serialize_usuario(item))


@csrf_exempt
@api_login_required
@require_http_methods(['POST'])
def usuario_foto(request, item_id):
    """Actualiza la foto de perfil del usuario."""
    item = get_object_or_404(Usuario, id=item_id)
    return JsonResponse(serialize_usuario(item))


# ============================================================
# API FICHAS
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def fichas_collection(request):
    """GET lista fichas, POST crea una ficha."""
    if request.method == 'GET':
        queryset = Ficha.objects.all()
        for item in queryset:
            sincronizar_trimestres_ficha(item)
        return JsonResponse([serialize_ficha(item) for item in queryset], safe=False)
    data = parse_json(request)
    item = Ficha.objects.create(
        codigo_ficha=normalize_text(data.get('codigoFicha')),
        programa_formacion=normalize_text(data.get('programaFormacion')),
        nivel=choice_to_db(data.get('nivel')) or Ficha.Nivel.TECNICO,
        jornada=choice_to_db(data.get('jornada')),
        modalidad=choice_to_db(data.get('modalidad')),
        fecha_inicio=normalize_date(data.get('fechaInicio')),
        fecha_fin=normalize_date(data.get('fechaFin')),
        estado=choice_to_db(data.get('estado')) or Ficha.Estado.ACTIVA,
    )
    sincronizar_trimestres_ficha(item)
    return JsonResponse(serialize_ficha(item), status=201)


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def fichas_detail(request, item_id):
    """GET, PUT o DELETE sobre una ficha especifica."""
    item = get_object_or_404(Ficha, id=item_id)
    if request.method == 'GET':
        sincronizar_trimestres_ficha(item)
        return JsonResponse(serialize_ficha(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    item.codigo_ficha = normalize_text(data.get('codigoFicha')) or item.codigo_ficha
    item.programa_formacion = normalize_text(data.get('programaFormacion')) or item.programa_formacion
    item.nivel = choice_to_db(data.get('nivel')) or item.nivel
    item.jornada = choice_to_db(data.get('jornada')) or item.jornada
    item.modalidad = choice_to_db(data.get('modalidad')) or item.modalidad
    item.fecha_inicio = normalize_date(data.get('fechaInicio')) or item.fecha_inicio
    item.fecha_fin = normalize_date(data.get('fechaFin')) or item.fecha_fin
    item.estado = choice_to_db(data.get('estado')) or item.estado
    item.save()
    sincronizar_trimestres_ficha(item)
    return JsonResponse(serialize_ficha(item))


# ============================================================
# API TRIMESTRES
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def trimestres_collection(request):
    """GET lista trimestres, POST crea un trimestre."""
    if request.method == 'GET':
        for ficha in Ficha.objects.all():
            sincronizar_trimestres_ficha(ficha)
        queryset = Trimestre.objects.select_related('ficha').all()
        return JsonResponse([serialize_trimestre(item) for item in queryset], safe=False)
    data = parse_json(request)
    item = Trimestre.objects.create(
        numero=data.get('numero'),
        ficha=get_object_or_404(Ficha, id=data.get('fichaId')),
        fecha_inicio=normalize_date(data.get('fechaInicio')),
        fecha_fin=normalize_date(data.get('fechaFin')),
        estado=choice_to_db(data.get('estado')) or Trimestre.Estado.ACTIVO,
    )
    return JsonResponse(serialize_trimestre(item), status=201)


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def trimestres_detail(request, item_id):
    """GET, PUT o DELETE sobre un trimestre especifico."""
    item = get_object_or_404(Trimestre.objects.select_related('ficha'), id=item_id)
    if request.method == 'GET':
        return JsonResponse(serialize_trimestre(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    item.numero = data.get('numero') or item.numero
    item.ficha = get_fk(Ficha, data.get('fichaId')) or item.ficha
    item.fecha_inicio = normalize_date(data.get('fechaInicio')) or item.fecha_inicio
    item.fecha_fin = normalize_date(data.get('fechaFin')) or item.fecha_fin
    item.estado = choice_to_db(data.get('estado')) or item.estado
    item.save()
    return JsonResponse(serialize_trimestre(item))


# ============================================================
# API GAES
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def gaes_collection(request):
    """GET lista GAES y POST crea un GAES nuevo."""
    if request.method == 'GET':
        queryset = Gaes.objects.select_related('ficha')
        ficha_id = request.GET.get('fichaId')
        if ficha_id:
            queryset = queryset.filter(ficha_id=ficha_id)
        return JsonResponse([serialize_gaes(item, include_integrantes=True) for item in queryset], safe=False)
    data = parse_json(request)
    item = Gaes.objects.create(
        nombre=normalize_text(data.get('nombre')),
        ficha=get_object_or_404(Ficha, id=data.get('fichaId')),
    )
    return JsonResponse(serialize_gaes(item, include_integrantes=True), status=201)


@api_login_required
def gaes_con_integrantes_todos(request):
    """Devuelve todos los GAES incluyendo sus integrantes."""
    queryset = Gaes.objects.select_related('ficha')
    return JsonResponse([serialize_gaes(item, include_integrantes=True) for item in queryset], safe=False)


@api_login_required
def gaes_con_integrantes(request, item_id):
    """Devuelve un GAES especifico con sus integrantes."""
    item = get_object_or_404(Gaes.objects.select_related('ficha'), id=item_id)
    return JsonResponse(serialize_gaes(item, include_integrantes=True))


@api_login_required
def gaes_por_ficha_con_integrantes(request, ficha_id):
    """Devuelve los GAES de una ficha incluyendo sus integrantes."""
    queryset = Gaes.objects.select_related('ficha').filter(ficha_id=ficha_id)
    return JsonResponse([serialize_gaes(item, include_integrantes=True) for item in queryset], safe=False)


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def gaes_detail(request, item_id):
    """GET, PUT o DELETE sobre un GAES especifico."""
    item = get_object_or_404(Gaes, id=item_id)
    if request.method == 'GET':
        return JsonResponse(serialize_gaes(item, include_integrantes=True))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    item.nombre = normalize_text(data.get('nombre')) or item.nombre
    item.ficha = get_fk(Ficha, data.get('fichaId')) or item.ficha
    item.save()
    return JsonResponse(serialize_gaes(item, include_integrantes=True))


# ============================================================
# API APRENDICES
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def aprendices_collection(request):
    """GET lista aprendices, POST crea un aprendiz."""
    if request.method == 'GET':
        queryset = Aprendiz.objects.select_related('usuario', 'ficha')
        return JsonResponse([serialize_aprendiz(item) for item in queryset], safe=False)
    data = parse_json(request)
    usuario = get_object_or_404(Usuario, id=data.get('usuarioId'))
    ficha = get_fk(Ficha, data.get('fichaId'))
    validation_error = validate_aprendiz_en_ficha(usuario, ficha)
    if validation_error:
        return JsonResponse({'detail': validation_error}, status=400)
    item = Aprendiz.objects.create(
        usuario=usuario,
        ficha=ficha,
    )
    if data.get('gaesId'):
        AprendizGaes.objects.get_or_create(aprendiz=item, gaes=get_object_or_404(Gaes, id=data.get('gaesId')))
    return JsonResponse(serialize_aprendiz(item), status=201)


@csrf_exempt
@api_login_required
@require_http_methods(['POST'])
def aprendices_carga_masiva(request):
    """Carga masiva de aprendices para una sola ficha a partir de un JSON."""
    ensure_default_roles()
    data = parse_json(request)
    ficha_id = data.get('fichaId')
    ficha = get_fk(Ficha, ficha_id)

    # La carga masiva siempre necesita una ficha destino valida.
    if ficha is None:
        return JsonResponse({'detail': 'Debes seleccionar una ficha valida'}, status=400)

    aprendices_json = data.get('aprendices')
    if not isinstance(aprendices_json, list) or len(aprendices_json) == 0:
        return JsonResponse({'detail': 'El JSON debe incluir una lista de aprendices'}, status=400)

    # Antes de procesar, se valida que la ficha no supere el limite de 30 aprendices.
    actuales = Aprendiz.objects.filter(ficha_id=ficha.id).count()
    if actuales + len(aprendices_json) > 30:
        cupos_disponibles = max(30 - actuales, 0)
        return JsonResponse(
            {
                'detail': (
                    f'La ficha {ficha.codigo_ficha} ya tiene {actuales} aprendices, '
                    f'solo permite maximo 30 y ahora mismo le quedan {cupos_disponibles} cupos'
                ),
            },
            status=400,
        )

    errores = []
    correos_en_json = set()
    documentos_en_json = set()
    aprendices_normalizados = []

    # Aqui se valida fila por fila el JSON recibido antes de guardar algo en la base de datos.
    for index, aprendiz in enumerate(aprendices_json, start=1):
        if not isinstance(aprendiz, dict):
            errores.append({'fila': index, 'mensaje': 'Cada aprendiz debe ser un objeto JSON valido'})
            continue

        nombre = normalize_text(aprendiz.get('nombre'))
        apellido = normalize_text(aprendiz.get('apellido'))
        correo = normalize_text(aprendiz.get('correo')).lower()
        tipo_documento = normalize_text(aprendiz.get('tipoDocumento'))
        numero_documento = normalize_text(aprendiz.get('numeroDocumento'))

        faltantes = []
        if not nombre:
            faltantes.append('nombre')
        if not apellido:
            faltantes.append('apellido')
        if not correo:
            faltantes.append('correo')
        if not tipo_documento:
            faltantes.append('tipoDocumento')
        if not numero_documento:
            faltantes.append('numeroDocumento')

        if faltantes:
            errores.append({'fila': index, 'mensaje': f'Faltan campos obligatorios: {", ".join(faltantes)}'})
            continue

        if correo in correos_en_json:
            errores.append({'fila': index, 'mensaje': f'El correo {correo} esta repetido dentro del JSON'})
            continue
        correos_en_json.add(correo)

        if numero_documento in documentos_en_json:
            errores.append({'fila': index, 'mensaje': f'El documento {numero_documento} esta repetido dentro del JSON'})
            continue
        documentos_en_json.add(numero_documento)

        validation_error = validate_unique_usuario_fields(correo, numero_documento)
        if validation_error:
            errores.append({'fila': index, 'mensaje': validation_error})
            continue

        if Aprendiz.objects.filter(
            ficha=ficha,
            usuario__nombre__iexact=nombre,
            usuario__apellido__iexact=apellido,
            usuario__numero_documento=numero_documento,
        ).exists():
            errores.append({'fila': index, 'mensaje': 'Ese aprendiz ya existe en la ficha seleccionada'})
            continue

        aprendices_normalizados.append(
            {
                'fila': index,
                'nombre': nombre,
                'apellido': apellido,
                'correo': correo,
                'tipo_documento': tipo_documento,
                'numero_documento': numero_documento,
            }
        )

    # Si existe cualquier error, no se guarda ningun aprendiz.
    if errores:
        return JsonResponse(
            {
                'message': 'La carga masiva contiene errores y no se guardo ningun aprendiz',
                'errores': errores,
                'totalRecibidos': len(aprendices_json),
                'totalValidos': len(aprendices_normalizados),
            },
            status=400,
        )

    rol_aprendiz = Rol.objects.filter(nombre_rol__iexact='aprendiz').first() or get_object_or_404(Rol, id=3)
    creados = []

    # La creacion se hace dentro de una transaccion para que todo quede consistente.
    with transaction.atomic():
        for aprendiz in aprendices_normalizados:
            usuario = Usuario.objects.create(
                nombre=aprendiz['nombre'],
                apellido=aprendiz['apellido'],
                correo=aprendiz['correo'],
                tipo_documento=aprendiz['tipo_documento'],
                numero_documento=aprendiz['numero_documento'],
                estado=Usuario.Estado.ACTIVO,
                rol=rol_aprendiz,
                password=make_password(aprendiz['numero_documento']),
                debe_cambiar_password=False,
                password_temporal=False,
            )
            perfil = Aprendiz.objects.create(usuario=usuario, ficha=ficha)
            creados.append(
                {
                    'id': perfil.id,
                    'usuarioId': usuario.id,
                    'nombre': usuario.nombre,
                    'apellido': usuario.apellido,
                    'correo': usuario.correo,
                    'numeroDocumento': usuario.numero_documento,
                    'passwordInicial': usuario.numero_documento,
                }
            )

    return JsonResponse(
        {
            'message': f'Se cargaron {len(creados)} aprendices en la ficha {ficha.codigo_ficha}',
            'ficha': serialize_ficha(ficha),
            'creados': creados,
            'totalCreados': len(creados),
        },
        status=201,
    )


@csrf_exempt
@api_login_required
@require_http_methods(['POST'])
def instructores_carga_masiva(request):
    """Carga masiva de instructores para una ficha a partir de un JSON."""
    ensure_default_roles()
    data = parse_json(request)
    ficha_id = data.get('fichaId')
    ficha = get_fk(Ficha, ficha_id)

    # La carga masiva siempre necesita una ficha destino valida.
    if ficha is None:
        return JsonResponse({'detail': 'Debes seleccionar una ficha valida'}, status=400)

    instructores_json = data.get('instructores')
    if not isinstance(instructores_json, list) or len(instructores_json) == 0:
        return JsonResponse({'detail': 'El JSON debe incluir una lista de instructores'}, status=400)

    errores = []
    correos_en_json = set()
    documentos_en_json = set()
    instructores_normalizados = []

    # Aqui se valida fila por fila el JSON recibido antes de guardar algo en la base de datos.
    for index, instructor in enumerate(instructores_json, start=1):
        if not isinstance(instructor, dict):
            errores.append({'fila': index, 'mensaje': 'Cada instructor debe ser un objeto JSON valido'})
            continue

        nombre = normalize_text(instructor.get('nombre'))
        apellido = normalize_text(instructor.get('apellido'))
        correo = normalize_text(instructor.get('correo')).lower()
        tipo_documento = normalize_text(instructor.get('tipoDocumento'))
        numero_documento = normalize_text(instructor.get('numeroDocumento'))
        especialidad = normalize_text(instructor.get('especialidad'))

        faltantes = []
        if not nombre:
            faltantes.append('nombre')
        if not apellido:
            faltantes.append('apellido')
        if not correo:
            faltantes.append('correo')
        if not tipo_documento:
            faltantes.append('tipoDocumento')
        if not numero_documento:
            faltantes.append('numeroDocumento')

        if faltantes:
            errores.append({'fila': index, 'mensaje': f'Faltan campos obligatorios: {", ".join(faltantes)}'})
            continue

        if correo in correos_en_json:
            errores.append({'fila': index, 'mensaje': f'El correo {correo} esta repetido dentro del JSON'})
            continue
        correos_en_json.add(correo)

        if numero_documento in documentos_en_json:
            errores.append({'fila': index, 'mensaje': f'El documento {numero_documento} esta repetido dentro del JSON'})
            continue
        documentos_en_json.add(numero_documento)

        validation_error = validate_unique_usuario_fields(correo, numero_documento)
        if validation_error:
            errores.append({'fila': index, 'mensaje': validation_error})
            continue

        instructores_normalizados.append(
            {
                'fila': index,
                'nombre': nombre,
                'apellido': apellido,
                'correo': correo,
                'tipo_documento': tipo_documento,
                'numero_documento': numero_documento,
                'especialidad': especialidad,
            }
        )

    # Si existe cualquier error, no se guarda ningun instructor.
    if errores:
        return JsonResponse(
            {
                'message': 'La carga masiva contiene errores y no se guardo ningun instructor',
                'errores': errores,
                'totalRecibidos': len(instructores_json),
                'totalValidos': len(instructores_normalizados),
            },
            status=400,
        )

    rol_instructor = Rol.objects.filter(nombre_rol__iexact='instructor').first() or get_object_or_404(Rol, id=2)
    creados = []

    # La creacion se hace dentro de una transaccion para que todo quede consistente.
    with transaction.atomic():
        for instructor in instructores_normalizados:
            usuario = Usuario.objects.create(
                nombre=instructor['nombre'],
                apellido=instructor['apellido'],
                correo=instructor['correo'],
                tipo_documento=instructor['tipo_documento'],
                numero_documento=instructor['numero_documento'],
                estado=Usuario.Estado.ACTIVO,
                rol=rol_instructor,
                password=make_password(instructor['numero_documento']),
                debe_cambiar_password=True,
                password_temporal=True,
            )
            perfil = Instructor.objects.create(
                usuario=usuario,
                ficha=ficha,
                especialidad=instructor['especialidad'] or None,
            )
            creados.append(
                {
                    'id': perfil.id,
                    'usuarioId': usuario.id,
                    'nombre': usuario.nombre,
                    'apellido': usuario.apellido,
                    'correo': usuario.correo,
                    'numeroDocumento': usuario.numero_documento,
                    'especialidad': perfil.especialidad or '',
                    'passwordInicial': usuario.numero_documento,
                }
            )

    return JsonResponse(
        {
            'message': f'Se cargaron {len(creados)} instructores en la ficha {ficha.codigo_ficha}',
            'ficha': serialize_ficha(ficha),
            'creados': creados,
            'totalCreados': len(creados),
        },
        status=201,
    )


@csrf_exempt
@api_login_required
@require_http_methods(['POST'])
def aprendices_por_ids(request):
    """Crea un aprendiz desde el admin o busca aprendices por lista de ids."""
    data = parse_json(request)

    if data.get('usuarioId'):
        usuario = get_object_or_404(Usuario, id=data.get('usuarioId'))
        ficha = get_fk(Ficha, data.get('fichaId'))
        validation_error = validate_aprendiz_en_ficha(usuario, ficha)
        if validation_error:
            return JsonResponse({'detail': validation_error}, status=400)

        item = Aprendiz.objects.create(usuario=usuario, ficha=ficha)
        if data.get('gaesId'):
            AprendizGaes.objects.get_or_create(aprendiz=item, gaes=get_object_or_404(Gaes, id=data.get('gaesId')))
        return JsonResponse(serialize_aprendiz(item), status=201)

    ids = data.get('ids') or []
    queryset = Aprendiz.objects.filter(id__in=ids).select_related('usuario', 'ficha')
    return JsonResponse([serialize_aprendiz(item) for item in queryset], safe=False)


@api_login_required
def aprendiz_por_usuario_dto(request, usuario_id):
    """Devuelve el perfil aprendiz asociado a un usuario."""
    aprendiz = get_object_or_404(Aprendiz.objects.select_related('usuario', 'ficha'), usuario_id=usuario_id)
    return JsonResponse(serialize_aprendiz(aprendiz))


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def aprendices_detail(request, item_id):
    """GET, PUT o DELETE sobre un aprendiz especifico."""
    item = get_object_or_404(Aprendiz.objects.select_related('usuario', 'ficha'), id=item_id)
    if request.method == 'GET':
        return JsonResponse(serialize_aprendiz(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    nuevo_usuario = get_fk(Usuario, data.get('usuarioId')) or item.usuario
    nueva_ficha = get_fk(Ficha, data.get('fichaId')) if data.get('fichaId') is not None else item.ficha
    validation_error = validate_aprendiz_en_ficha(nuevo_usuario, nueva_ficha, exclude_aprendiz_id=item.id)
    if validation_error:
        return JsonResponse({'detail': validation_error}, status=400)
    item.usuario = nuevo_usuario
    item.ficha = nueva_ficha
    item.save()
    if data.get('gaesId') is not None:
        AprendizGaes.objects.filter(aprendiz=item).delete()
        if data.get('gaesId'):
            AprendizGaes.objects.get_or_create(aprendiz=item, gaes=get_object_or_404(Gaes, id=data.get('gaesId')))
    if data.get('estado'):
        item.usuario.estado = choice_to_db(data.get('estado'))
        item.usuario.save(update_fields=['estado'])
    return JsonResponse(serialize_aprendiz(item))


@csrf_exempt
@api_login_required
@require_http_methods(['PATCH'])
def aprendiz_activar(request, item_id):
    """Activa un aprendiz."""
    item = get_object_or_404(Aprendiz, id=item_id)
    item.usuario.estado = Usuario.Estado.ACTIVO
    item.usuario.save(update_fields=['estado'])
    return JsonResponse(serialize_aprendiz(item))


@csrf_exempt
@api_login_required
@require_http_methods(['PATCH'])
def aprendiz_desactivar(request, item_id):
    """Desactiva un aprendiz."""
    item = get_object_or_404(Aprendiz, id=item_id)
    item.usuario.estado = Usuario.Estado.INACTIVO
    item.usuario.save(update_fields=['estado'])
    return JsonResponse(serialize_aprendiz(item))


@csrf_exempt
@api_login_required
@require_http_methods(['PUT', 'DELETE'])
def aprendiz_asignar_gaes(request, item_id, gaes_id):
    """Asigna un aprendiz a un GAES o lo remueve si llega DELETE."""
    item = get_object_or_404(Aprendiz, id=item_id)
    if request.method == 'DELETE':
        AprendizGaes.objects.filter(aprendiz=item, gaes_id=gaes_id).delete()
        return JsonResponse(serialize_aprendiz(item))

    gaes = get_object_or_404(Gaes, id=gaes_id)
    AprendizGaes.objects.filter(aprendiz=item).delete()
    AprendizGaes.objects.get_or_create(aprendiz=item, gaes=gaes)
    return JsonResponse(serialize_aprendiz(item))
    return JsonResponse(serialize_aprendiz(item))


@csrf_exempt
@api_login_required
@require_http_methods(['DELETE'])
def aprendiz_quitar_gaes(request, item_id):
    """Quita la relacion de un aprendiz con su GAES."""
    item = get_object_or_404(Aprendiz, id=item_id)
    AprendizGaes.objects.filter(aprendiz=item).delete()
    return JsonResponse(serialize_aprendiz(item))


@csrf_exempt
@api_login_required
@require_http_methods(['PUT'])
def gaes_asignar_aprendices(request, item_id):
    """Asigna hasta 5 aprendices de la misma ficha a un GAES."""
    gaes = get_object_or_404(Gaes, id=item_id)
    data = parse_json(request)
    integrantes = data.get('integrantes') or []

    aprendiz_ids = []
    for integrante in integrantes:
        integrante_id = integrante.get('id') if isinstance(integrante, dict) else integrante
        if integrante_id:
            aprendiz_ids.append(int(integrante_id))

    aprendiz_ids = list(dict.fromkeys(aprendiz_ids))

    if len(aprendiz_ids) > 5:
        return JsonResponse({'detail': 'Cada GAES puede tener maximo 5 aprendices'}, status=400)

    aprendices = list(Aprendiz.objects.select_related('usuario', 'ficha').filter(id__in=aprendiz_ids))
    if len(aprendices) != len(aprendiz_ids):
        return JsonResponse({'detail': 'Uno o mas aprendices no existen'}, status=400)

    invalidos = [aprendiz.id for aprendiz in aprendices if aprendiz.ficha_id != gaes.ficha_id]
    if invalidos:
        return JsonResponse({'detail': 'Solo puedes asignar aprendices de la misma ficha del GAES'}, status=400)

    AprendizGaes.objects.filter(gaes=gaes).exclude(aprendiz_id__in=aprendiz_ids).delete()

    for aprendiz in aprendices:
        AprendizGaes.objects.filter(aprendiz=aprendiz).exclude(gaes=gaes).delete()
        AprendizGaes.objects.get_or_create(aprendiz=aprendiz, gaes=gaes)

    return JsonResponse(serialize_gaes(gaes, include_integrantes=True))


@api_login_required
def aprendices_por_ficha_dto(request, ficha_id):
    """Lista aprendices de una ficha con datos utiles para asignacion a GAES."""
    queryset = Aprendiz.objects.select_related('usuario', 'ficha').filter(ficha_id=ficha_id)
    return JsonResponse([serialize_aprendiz(item) for item in queryset], safe=False)


# ============================================================
# API INSTRUCTORES
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def instructores_collection(request):
    """GET lista instructores, POST crea un instructor."""
    if request.method == 'GET':
        queryset = Instructor.objects.select_related('usuario', 'ficha')
        usuario_id = request.GET.get('usuarioId')
        if usuario_id:
            queryset = queryset.filter(usuario_id=usuario_id)
        return JsonResponse([serialize_instructor(item) for item in queryset], safe=False)
    data = parse_json(request)
    item = Instructor.objects.create(
        usuario=get_object_or_404(Usuario, id=data.get('usuarioId')),
        ficha=get_fk(Ficha, data.get('fichaId')),
        especialidad=normalize_text(data.get('especialidad')),
    )
    return JsonResponse(serialize_instructor(item), status=201)


@api_login_required
def instructor_por_usuario(request, usuario_id):
    """Devuelve el perfil instructor asociado a un usuario."""
    item = get_preferred_instructor_by_user_id(usuario_id)
    if item is None:
        return JsonResponse({'detail': 'Instructor no encontrado'}, status=404)
    return JsonResponse(serialize_instructor(item))


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def instructores_detail(request, item_id):
    """GET, PUT o DELETE sobre un instructor especifico."""
    item = get_object_or_404(Instructor.objects.select_related('usuario', 'ficha'), id=item_id)
    if request.method == 'GET':
        return JsonResponse(serialize_instructor(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    item.usuario = get_fk(Usuario, data.get('usuarioId')) or item.usuario
    if data.get('fichaId') is not None:
        item.ficha = get_fk(Ficha, data.get('fichaId'))
    item.especialidad = normalize_text(data.get('especialidad')) or item.especialidad
    item.save()
    if data.get('estado'):
        item.usuario.estado = choice_to_db(data.get('estado'))
        item.usuario.save(update_fields=['estado'])
    return JsonResponse(serialize_instructor(item))


# ============================================================
# API PROYECTOS
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def proyectos_collection(request):
    """GET lista proyectos, POST crea un proyecto."""
    if request.method == 'GET':
        queryset = Proyecto.objects.select_related('gaes')
        return JsonResponse([serialize_proyecto(item) for item in queryset], safe=False)

    data = parse_json(request)
    item = Proyecto.objects.create(
        nombre=normalize_text(data.get('nombre')),
        descripcion=normalize_text(data.get('descripcion')),
        gaes=get_object_or_404(Gaes, id=data.get('gaesId')),
        estado=proyecto_estado_to_db(data.get('estado')),
        fecha_inicio=normalize_date(data.get('fechaInicio')),
        fecha_fin=normalize_date(data.get('fechaFin')),
    )
    return JsonResponse(serialize_proyecto(item), status=201)


@api_login_required
def proyectos_por_lider(request, aprendiz_id):
    """Lista proyectos del o los GAES a los que pertenece el aprendiz indicado."""
    gaes_ids = AprendizGaes.objects.filter(aprendiz_id=aprendiz_id).values_list('gaes_id', flat=True)
    queryset = Proyecto.objects.select_related('gaes').filter(gaes_id__in=gaes_ids)
    return JsonResponse([serialize_proyecto(item) for item in queryset], safe=False)


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def proyectos_detail(request, item_id):
    """GET, PUT o DELETE sobre un proyecto especifico."""
    item = get_object_or_404(Proyecto.objects.select_related('gaes'), id=item_id)
    if request.method == 'GET':
        return JsonResponse(serialize_proyecto(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)

    data = parse_json(request)
    item.nombre = normalize_text(data.get('nombre')) or item.nombre
    item.descripcion = normalize_text(data.get('descripcion')) or item.descripcion
    item.gaes = get_fk(Gaes, data.get('gaesId')) or item.gaes
    item.estado = proyecto_estado_to_db(data.get('estado')) or item.estado
    item.fecha_inicio = normalize_date(data.get('fechaInicio')) or item.fecha_inicio
    item.fecha_fin = normalize_date(data.get('fechaFin')) or item.fecha_fin
    item.save()
    return JsonResponse(serialize_proyecto(item))


# ============================================================
# API EVALUACIONES
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def evaluaciones_collection(request):
    """GET lista evaluaciones, POST crea una evaluacion."""
    if request.method == 'GET':
        queryset = Evaluacion.objects.select_related('entregable', 'aprendiz__usuario', 'aprendiz__ficha', 'gaes', 'evaluador__usuario')
        usuario_id = request.GET.get('usuarioId')
        if usuario_id:
            aprendiz = Aprendiz.objects.filter(usuario_id=usuario_id).first()
            queryset = queryset.filter(aprendiz=aprendiz) if aprendiz else queryset.none()
        return JsonResponse([serialize_evaluacion(item) for item in queryset], safe=False)
    data = parse_json(request)

    entregable = get_fk(Entregable, get_payload_id(data, 'entregableId'))
    proyecto_id = get_payload_id(data, 'proyectoId')
    aprendiz_id = get_payload_id(data, 'aprendizId')
    gaes_id = get_payload_id(data, 'gaesId')
    evaluador_id = get_payload_id(data, 'evaluadorId') or get_payload_id(data, 'instructorId')

    if entregable is None and proyecto_id and aprendiz_id:
        entregable = (
            Entregable.objects.filter(proyecto_id=proyecto_id, aprendiz_id=aprendiz_id)
            .order_by('-id')
            .first()
        )

    if entregable is None and aprendiz_id:
        entregable = (
            Entregable.objects.filter(aprendiz_id=aprendiz_id)
            .order_by('-id')
            .first()
        )

    if entregable is None and gaes_id:
        entregable = (
            Entregable.objects.filter(proyecto__gaes_id=gaes_id)
            .order_by('-id')
            .first()
        )

    if entregable is None and proyecto_id:
        entregable = Entregable.objects.filter(proyecto_id=proyecto_id).order_by('-id').first()

    if entregable is None:
        return JsonResponse({'detail': 'No se encontro entregable para evaluar'}, status=400)

    item = Evaluacion.objects.create(
        entregable=entregable,
        aprendiz=get_fk(Aprendiz, aprendiz_id),
        gaes=get_fk(Gaes, gaes_id),
        evaluador=get_object_or_404(Instructor, id=evaluador_id),
        calificacion=data.get('calificacion') or None,
        observaciones=normalize_text(data.get('observaciones')),
        fecha=normalize_date(data.get('fecha')) or timezone.now().date(),
    )
    return JsonResponse(serialize_evaluacion(item), status=201)


@api_login_required
def evaluaciones_por_aprendiz(request, aprendiz_id):
    """Lista evaluaciones asociadas a un aprendiz."""
    queryset = Evaluacion.objects.filter(aprendiz_id=aprendiz_id).select_related('entregable', 'aprendiz__usuario', 'aprendiz__ficha', 'gaes', 'evaluador__usuario')
    return JsonResponse([serialize_evaluacion(item) for item in queryset], safe=False)


@csrf_exempt
@api_login_required
@require_http_methods(['POST'])
def evaluaciones_por_gaes(request):
    """Lista o crea evaluaciones filtradas por GAES."""
    data = parse_json(request)

    if data.get('calificacion') is not None:
        gaes_id = get_payload_id(data, 'gaesId')
        evaluador_id = get_payload_id(data, 'evaluadorId') or get_payload_id(data, 'instructorId')
        gaes = get_object_or_404(Gaes, id=gaes_id)
        evaluador = get_object_or_404(Instructor, id=evaluador_id)
        aprendices_ids = list(AprendizGaes.objects.filter(gaes=gaes).values_list('aprendiz_id', flat=True))

        if not aprendices_ids:
            return JsonResponse({'detail': 'El GAES no tiene aprendices para calificar'}, status=400)

        creadas = []
        for aprendiz_id in aprendices_ids:
            entregable = (
                Entregable.objects.filter(aprendiz_id=aprendiz_id, proyecto__gaes_id=gaes.id)
                .order_by('-id')
                .first()
            ) or (
                Entregable.objects.filter(proyecto__gaes_id=gaes.id)
                .order_by('-id')
                .first()
            )

            if not entregable:
                continue

            item = Evaluacion.objects.create(
                entregable=entregable,
                aprendiz_id=aprendiz_id,
                gaes=gaes,
                evaluador=evaluador,
                calificacion=data.get('calificacion') or None,
                observaciones=normalize_text(data.get('observaciones')),
                fecha=normalize_date(data.get('fecha')) or timezone.now().date(),
            )
            creadas.append(serialize_evaluacion(item))

        if not creadas:
            return JsonResponse({'detail': 'No se encontro entregable para los aprendices del GAES'}, status=400)

        return JsonResponse(creadas, safe=False, status=201)

    queryset = Evaluacion.objects.filter(gaes_id=data.get('gaesId')).select_related('entregable', 'aprendiz__usuario', 'aprendiz__ficha', 'gaes', 'evaluador__usuario')
    return JsonResponse([serialize_evaluacion(item) for item in queryset], safe=False)


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def evaluaciones_detail(request, item_id):
    """GET, PUT o DELETE sobre una evaluacion especifica."""
    item = get_object_or_404(Evaluacion.objects.select_related('entregable', 'aprendiz__usuario', 'aprendiz__ficha', 'gaes', 'evaluador__usuario'), id=item_id)
    if request.method == 'GET':
        return JsonResponse(serialize_evaluacion(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    item.entregable = get_fk(Entregable, data.get('entregableId')) or item.entregable
    item.aprendiz = get_fk(Aprendiz, data.get('aprendizId')) or item.aprendiz
    item.gaes = get_fk(Gaes, data.get('gaesId')) if data.get('gaesId') is not None else item.gaes
    item.evaluador = get_fk(Instructor, data.get('evaluadorId')) or item.evaluador
    item.calificacion = data.get('calificacion') if data.get('calificacion') is not None else item.calificacion
    item.observaciones = normalize_text(data.get('observaciones')) or item.observaciones
    item.fecha = normalize_date(data.get('fecha')) or item.fecha
    item.save()
    return JsonResponse(serialize_evaluacion(item))


@api_login_required
def evaluaciones_descargar_excel(request):
    """Exporta evaluaciones en CSV para descarga."""
    queryset = Evaluacion.objects.select_related('aprendiz__usuario', 'gaes', 'evaluador__usuario')
    rows = [
        [
            item.id,
            item.aprendiz.usuario.nombre if item.aprendiz_id else '',
            item.gaes.nombre if item.gaes_id else '',
            item.evaluador.usuario.nombre if item.evaluador_id else '',
            item.calificacion or '',
            item.fecha,
        ]
        for item in queryset
    ]
    return csv_response('evaluaciones.csv', ['ID', 'Aprendiz', 'GAES', 'Evaluador', 'Calificacion', 'Fecha'], rows)


# ============================================================
# API ENTREGABLES
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def entregables_collection(request):
    """GET lista entregables, POST crea un entregable."""
    if request.method == 'GET':
        queryset = Entregable.objects.select_related('proyecto', 'trimestre')
        return JsonResponse([serialize_entregable(item) for item in queryset], safe=False)

    if request.content_type and 'multipart/form-data' in request.content_type:
        data = request.POST
        archivo = request.FILES.get('archivo')
    else:
        data = parse_json(request)
        archivo = None

    item = Entregable.objects.create(
        nombre=normalize_text(data.get('nombre')),
        descripcion=normalize_text(data.get('descripcion')),
        proyecto=get_fk(Proyecto, data.get('proyectoId')),
        trimestre=get_fk(Trimestre, data.get('trimestreId')),
        aprendiz=get_fk(Aprendiz, data.get('aprendizId')),
        url=data.get('documentoUrl') or data.get('url') or None,
        archivo=archivo.read() if archivo else None,
        nombre_archivo=archivo.name if archivo else (data.get('nombreArchivo') or None),
    )
    return JsonResponse(serialize_entregable(item), status=201)


@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'PUT', 'DELETE'])
def entregables_detail(request, item_id):
    """GET, PUT o DELETE sobre un entregable especifico."""
    item = get_object_or_404(Entregable, id=item_id)
    usuario = get_request_user(request)
    if usuario and usuario.rol and usuario.rol.nombre_rol.lower() == 'aprendiz':
        if not item.aprendiz_id or not item.aprendiz or item.aprendiz.usuario_id != usuario.id:
            return JsonResponse({'detail': 'No tienes permisos para editar este entregable'}, status=403)
    if request.method == 'GET':
        return JsonResponse(serialize_entregable(item))
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    if request.content_type and 'multipart/form-data' in request.content_type:
        data = request.POST
        archivo = request.FILES.get('archivo')
    else:
        data = parse_json(request)
        archivo = None
    item.nombre = normalize_text(data.get('nombre')) or item.nombre
    item.descripcion = normalize_text(data.get('descripcion')) or item.descripcion
    item.proyecto = get_fk(Proyecto, data.get('proyectoId')) if data.get('proyectoId') is not None else item.proyecto
    item.trimestre = get_fk(Trimestre, data.get('trimestreId')) if data.get('trimestreId') is not None else item.trimestre
    item.aprendiz = get_fk(Aprendiz, data.get('aprendizId')) if data.get('aprendizId') is not None else item.aprendiz
    if archivo:
        item.archivo = archivo.read()
        item.nombre_archivo = archivo.name
        item.url = None
    else:
        item.url = data.get('documentoUrl') or data.get('url') or item.url
        item.nombre_archivo = data.get('nombreArchivo') or item.nombre_archivo
    item.save()
    return JsonResponse(serialize_entregable(item))


@api_login_required
@require_http_methods(['GET'])
def entregables_descargar(request, item_id):
    """Descarga el archivo de un entregable o redirige a su URL."""
    item = get_object_or_404(Entregable, id=item_id)

    if item.archivo:
        response = HttpResponse(item.archivo, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{item.nombre_archivo or item.nombre or "entregable.bin"}"'
        return response

    if item.url:
        return JsonResponse({'url': item.url})

    return JsonResponse({'detail': 'Este entregable no tiene archivo para descargar'}, status=404)


# ============================================================
# API ADMINISTRADORES
# ============================================================

@csrf_exempt
@api_login_required
@require_http_methods(['GET', 'POST'])
def administradores_collection(request):
    """GET lista administradores, POST crea uno nuevo."""
    if request.method == 'GET':
        queryset = Administrador.objects.select_related('usuario__rol')
        return JsonResponse([serialize_administrador(item) for item in queryset], safe=False)
    data = parse_json(request)
    item = Administrador.objects.create(usuario=get_object_or_404(Usuario, id=(data.get('usuario') or {}).get('id')))
    return JsonResponse(serialize_administrador(item), status=201)


@csrf_exempt
@api_login_required
@require_http_methods(['PUT', 'DELETE'])
def administradores_detail(request, item_id):
    """PUT actualiza administrador, DELETE lo elimina."""
    item = get_object_or_404(Administrador.objects.select_related('usuario__rol'), id=item_id)
    if request.method == 'DELETE':
        item.delete()
        return HttpResponse(status=204)
    data = parse_json(request)
    item.usuario = get_object_or_404(Usuario, id=(data.get('usuario') or {}).get('id'))
    item.save()
    return JsonResponse(serialize_administrador(item))


# ============================================================
# REPORTES Y EXPORTACIONES
# ============================================================

def reporte_usuarios_general(request):
    """Genera reporte general de usuarios para el dashboard admin."""
    rows = Usuario.objects.select_related('rol').all()
    return JsonResponse(
        [
            {
                'id': item.id,
                'nombre': item.nombre,
                'apellido': item.apellido,
                'correo': item.correo,
                'tipoDocumento': item.tipo_documento,
                'numeroDocumento': item.numero_documento,
                'rol': item.rol.nombre_rol.upper() if item.rol else '',
                'estado': choice_to_api(item.estado),
            }
            for item in rows
        ],
        safe=False,
    )


def reporte_fichas_aprendices(request):
    """Genera reporte de fichas con total de aprendices."""
    rows = Aprendiz.objects.select_related('usuario', 'ficha').all()
    return JsonResponse(
        [
            {
                'id': item.id,
                'codigo_ficha': item.ficha.codigo_ficha if item.ficha_id else '',
                'programa_formacion': item.ficha.programa_formacion if item.ficha_id else '',
                'jornada': choice_to_api(item.ficha.jornada) if item.ficha_id else '',
                'modalidad': choice_to_api(item.ficha.modalidad) if item.ficha_id else '',
                'aprendiz': f'{item.usuario.nombre} {item.usuario.apellido}'.strip(),
                'tipo_documento': item.usuario.tipo_documento,
                'numero_documento': item.usuario.numero_documento,
                'estado_aprendiz': choice_to_api(item.usuario.estado),
                'estado_ficha': choice_to_api(item.ficha.estado) if item.ficha_id else '',
            }
            for item in rows
        ],
        safe=False,
    )


def reporte_instructores_especialidad(request):
    """Genera reporte detallado de instructores."""
    rows = Instructor.objects.select_related('usuario', 'ficha').all()
    return JsonResponse(
        [
            {
                'id': item.id,
                'instructor': f'{item.usuario.nombre} {item.usuario.apellido}'.strip(),
                'correo': item.usuario.correo,
                'ficha': item.ficha.codigo_ficha if item.ficha_id else '',
                'especialidad': item.especialidad or '',
                'estado': choice_to_api(item.usuario.estado),
            }
            for item in rows
        ],
        safe=False,
    )


def reporte_trimestres_ficha(request):
    """Genera reporte de trimestres ligados a fichas."""
    rows = Trimestre.objects.select_related('ficha').all()
    return JsonResponse(
        [
            {
                'id': item.id,
                'codigo_ficha': item.ficha.codigo_ficha if item.ficha_id else '',
                'trimestre': item.numero,
                'fecha_inicio': item.fecha_inicio.isoformat() if item.fecha_inicio else None,
                'fecha_fin': item.fecha_fin.isoformat() if item.fecha_fin else None,
                'estado': choice_to_api(item.estado),
            }
            for item in rows
        ],
        safe=False,
    )


def reporte_proyectos_gaes(request):
    """Genera reporte de proyectos por GAES."""
    rows = Proyecto.objects.select_related('gaes').all()
    return JsonResponse(
        [
            {
                'id': item.id,
                'gaes': item.gaes.nombre if item.gaes_id else '',
                'proyecto': item.nombre,
                'estado': choice_to_api(item.estado),
                'fecha_inicio': item.fecha_inicio.isoformat() if item.fecha_inicio else None,
                'fecha_fin': item.fecha_fin.isoformat() if item.fecha_fin else None,
            }
            for item in rows
        ],
        safe=False,
    )


@csrf_exempt
def descargar_reporte_usuarios_excel(request):
    """Descarga reporte de usuarios en Excel real."""
    rows = [
        [
            item.id,
            item.nombre,
            item.apellido,
            item.correo,
            item.tipo_documento,
            item.numero_documento,
            item.rol.nombre_rol.upper() if item.rol else '',
            choice_to_api(item.estado),
        ]
        for item in Usuario.objects.select_related('rol').all()
    ]
    return excel_response(
        'reporte-usuarios.xlsx',
        ['ID', 'Nombre', 'Apellido', 'Correo', 'Tipo Documento', 'Numero Documento', 'Rol', 'Estado'],
        rows,
        sheet_name='Usuarios',
    )


@csrf_exempt
def descargar_reporte_generico(request, reporte, formato):
    """Descarga reportes simples usando nombre de reporte y formato."""
    mapping = {
        'usuarios': (
            ['ID', 'Nombre', 'Apellido', 'Correo', 'Tipo Documento', 'Numero Documento', 'Rol', 'Estado'],
            [
                [
                    u.id,
                    u.nombre,
                    u.apellido,
                    u.correo,
                    u.tipo_documento,
                    u.numero_documento,
                    u.rol.nombre_rol.upper() if u.rol else '',
                    choice_to_api(u.estado),
                ]
                for u in Usuario.objects.select_related('rol')
            ],
        ),
        'fichas': (
            ['Codigo Ficha', 'Programa', 'Jornada', 'Modalidad', 'Aprendiz', 'Documento', 'Estado Aprendiz', 'Estado Ficha'],
            [
                [
                    item.ficha.codigo_ficha if item.ficha_id else '',
                    item.ficha.programa_formacion if item.ficha_id else '',
                    choice_to_api(item.ficha.jornada) if item.ficha_id else '',
                    choice_to_api(item.ficha.modalidad) if item.ficha_id else '',
                    f'{item.usuario.nombre} {item.usuario.apellido}'.strip(),
                    f'{item.usuario.tipo_documento} {item.usuario.numero_documento}'.strip(),
                    choice_to_api(item.usuario.estado),
                    choice_to_api(item.ficha.estado) if item.ficha_id else '',
                ]
                for item in Aprendiz.objects.select_related('usuario', 'ficha').all()
            ],
        ),
        'instructores': (['ID', 'Instructor', 'Ficha', 'Especialidad'], [[i.id, i.usuario.nombre, i.ficha.codigo_ficha if i.ficha_id else '', i.especialidad] for i in Instructor.objects.select_related('usuario', 'ficha')]),
        'trimestres': (['ID', 'Numero', 'Ficha'], [[t.id, t.numero, t.ficha.codigo_ficha] for t in Trimestre.objects.select_related('ficha')]),
        'proyectos': (['ID', 'Proyecto', 'Estado'], [[p.id, p.nombre, p.estado] for p in Proyecto.objects.all()]),
    }
    headers, rows = mapping.get(reporte, (['Mensaje'], [[f'Reporte {reporte} no disponible']]))
    if formato == 'pdf':
        return pdf_response(f'{reporte}.pdf', headers, rows, title=f'Reporte {reporte.title()}')
    if formato == 'excel':
        return excel_response(f'reporte-{reporte}.xlsx', headers, rows, sheet_name=reporte.title())
    return csv_response(f'{reporte}.csv', headers, rows)
